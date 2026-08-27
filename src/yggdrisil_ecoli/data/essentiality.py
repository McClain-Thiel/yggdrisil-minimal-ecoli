"""Frozen experimental essentiality observations and condition summary."""

from __future__ import annotations

import os
import tempfile
from collections import Counter
from collections.abc import Iterable
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Literal

import pyarrow as pa
import pyarrow.parquet as pq

from yggdrisil_ecoli.constants import REFERENCE_ACCESSION, is_b_number
from yggdrisil_ecoli.data.errors import DataValidationError
from yggdrisil_ecoli.data.registry import GeneRegistry

EssentialityClass = Literal[
    "essential", "conditionally_essential", "nonessential", "ambiguous", "unknown"
]

OBSERVATION_SCHEMA = pa.schema(
    [
        pa.field("observation_id", pa.string(), nullable=False),
        pa.field("b_number", pa.string(), nullable=False),
        pa.field("gene_symbol_raw", pa.string()),
        pa.field("study_id", pa.string(), nullable=False),
        pa.field("assay_id", pa.string(), nullable=False),
        pa.field("source_call_raw", pa.string(), nullable=False),
        pa.field("normalized_call", pa.string(), nullable=False),
        pa.field("insertion_count", pa.int64(), nullable=False),
        pa.field("ipkm", pa.float64(), nullable=False),
        pa.field("end_corrected_insertion_count", pa.int64(), nullable=False),
        pa.field("ecipkm", pa.float64(), nullable=False),
        pa.field("effect_metric", pa.string(), nullable=False),
        pa.field("threshold_rule", pa.string(), nullable=False),
        pa.field("classification_basis", pa.string(), nullable=False),
        pa.field("strain_name", pa.string(), nullable=False),
        pa.field("reference_accession", pa.string(), nullable=False),
        pa.field("medium_name", pa.string(), nullable=False),
        pa.field("medium_recipe", pa.string(), nullable=False),
        pa.field("glucose_g_l", pa.float64()),
        pa.field("oxygenation", pa.string(), nullable=False),
        pa.field("temperature_c", pa.float64(), nullable=False),
        pa.field("perturbation_type", pa.string(), nullable=False),
        pa.field("coverage_status", pa.string(), nullable=False),
        pa.field("target_relevance_tier", pa.string(), nullable=False),
        pa.field("source_start", pa.int64(), nullable=False),
        pa.field("source_end", pa.int64(), nullable=False),
        pa.field("pec_call_raw", pa.string()),
        pa.field("gerdes_call_raw", pa.string()),
        pa.field("doi", pa.string(), nullable=False),
        pa.field("license_spdx", pa.string(), nullable=False),
    ]
)

SUMMARY_SCHEMA = pa.schema(
    [
        pa.field("b_number", pa.string(), nullable=False),
        pa.field("classification", pa.string(), nullable=False),
        pa.field("m9_call_raw", pa.string()),
        pa.field("lb_call_raw", pa.string()),
        pa.field("m9_ecipkm", pa.float64()),
        pa.field("lb_ecipkm", pa.float64()),
        pa.field("cross_condition_pattern", pa.string()),
        pa.field("evidence_conflict", pa.bool_(), nullable=False),
        pa.field("basis_observation_ids", pa.list_(pa.string()), nullable=False),
        pa.field("study_id", pa.string()),
    ]
)

_STUDY_ID = "choe2023_tnseq"
_DOI = "10.1128/msystems.00896-22"
_EXPECTED_SOURCE_COUNTS = {
    "source_rows": 4498,
    "protein_coding_nonpseudo_rows": 4140,
    "lb_essential": 422,
    "lb_nonessential": 3718,
    "m9_essential": 545,
    "m9_nonessential": 3595,
}


@dataclass(frozen=True, slots=True)
class EssentialityObservation:
    observation_id: str
    b_number: str
    gene_symbol_raw: str | None
    study_id: str
    assay_id: str
    source_call_raw: str
    normalized_call: str
    insertion_count: int
    ipkm: float
    end_corrected_insertion_count: int
    ecipkm: float
    effect_metric: str
    threshold_rule: str
    classification_basis: str
    strain_name: str
    reference_accession: str
    medium_name: str
    medium_recipe: str
    glucose_g_l: float | None
    oxygenation: str
    temperature_c: float
    perturbation_type: str
    coverage_status: str
    target_relevance_tier: str
    source_start: int
    source_end: int
    pec_call_raw: str | None
    gerdes_call_raw: str | None
    doi: str
    license_spdx: str


@dataclass(frozen=True, slots=True)
class EssentialitySummary:
    b_number: str
    classification: EssentialityClass
    m9_call_raw: str | None
    lb_call_raw: str | None
    m9_ecipkm: float | None
    lb_ecipkm: float | None
    cross_condition_pattern: str | None
    evidence_conflict: bool
    basis_observation_ids: tuple[str, ...]
    study_id: str | None

    def as_arrow_row(self) -> dict[str, object]:
        row = asdict(self)
        row["basis_observation_ids"] = list(self.basis_observation_ids)
        return row


@dataclass(frozen=True, slots=True)
class EssentialityImportReport:
    source_rows: int
    protein_coding_nonpseudo_rows: int
    source_call_counts: dict[str, int]
    mapped_source_genes: int
    unmapped_source_ids: tuple[str, ...]
    canonical_genes_without_measurement: tuple[str, ...]
    coordinate_mismatches: tuple[str, ...]
    summary_counts: dict[str, int]

    def as_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class ParsedEssentiality:
    observations: tuple[EssentialityObservation, ...]
    summaries: tuple[EssentialitySummary, ...]
    report: EssentialityImportReport


class EssentialityDataset:
    """Lookup layer shared by the scorer and optional detail tool."""

    def __init__(
        self,
        summaries: Iterable[EssentialitySummary],
        observations: Iterable[EssentialityObservation],
    ) -> None:
        self._summaries = {summary.b_number: summary for summary in summaries}
        by_gene: dict[str, list[EssentialityObservation]] = {}
        for observation in observations:
            by_gene.setdefault(observation.b_number, []).append(observation)
        self._observations = {
            b_number: tuple(sorted(rows, key=lambda row: row.assay_id))
            for b_number, rows in by_gene.items()
        }

    @classmethod
    def from_parquet(
        cls, summary_path: str | Path, observation_path: str | Path
    ) -> EssentialityDataset:
        summary_table = pq.read_table(summary_path, schema=SUMMARY_SCHEMA)
        observation_table = pq.read_table(observation_path, schema=OBSERVATION_SCHEMA)
        summaries = []
        for row in summary_table.to_pylist():
            row["basis_observation_ids"] = tuple(row["basis_observation_ids"] or ())
            summaries.append(EssentialitySummary(**row))
        observations = [
            EssentialityObservation(**row) for row in observation_table.to_pylist()
        ]
        return cls(summaries, observations)

    def summary(self, b_number: str) -> EssentialitySummary:
        if not is_b_number(b_number):
            raise DataValidationError(
                f"expected a canonical b-number, got {b_number!r}"
            )
        try:
            return self._summaries[b_number]
        except KeyError as exc:
            raise KeyError(
                f"gene is absent from essentiality summary: {b_number}"
            ) from exc

    def observations(self, b_number: str) -> tuple[EssentialityObservation, ...]:
        self.summary(b_number)
        return self._observations.get(b_number, ())


def parse_choe_workbook(
    path: str | Path,
    registry: GeneRegistry,
    *,
    expected_source_counts: dict[str, int] | None = _EXPECTED_SOURCE_COUNTS,
) -> ParsedEssentiality:
    """Parse checksum-pinned Choe 2023 Table S1 and preserve both conditions."""

    # Workbook parsing is a build-time optional dependency. Runtime scorers need
    # only the generated Parquet tables and can import without openpyxl installed.
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    if workbook.sheetnames != ["Table S1"]:
        raise DataValidationError(
            f"unexpected Choe workbook sheets: {workbook.sheetnames}"
        )
    worksheet = workbook["Table S1"]
    header_one = tuple(
        cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
    )
    header_two = tuple(
        cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))
    )
    if (
        header_one[0:10]
        != (
            "Gene",
            "Start",
            "End",
            "Length (nt)",
            "Strand",
            "Locus Tag",
            "CDS",
            "Pseudo",
            "PEC",
            "Gerdes",
        )
        or header_one[10] != "LB medium"
        or header_one[15] != "M9 glucose (0.2%) medium"
        or header_two[10:15]
        != ("Insertion", "IPKM", "ec Insertion", "ecIPKM", "Essentiality")
        or header_two[15:20]
        != ("Insertion", "IPKM", "ec Insertion", "ecIPKM", "Essentiality")
    ):
        raise DataValidationError("unexpected Choe Table S1 column contract")

    source_rows = list(worksheet.iter_rows(min_row=3, values_only=True))
    protein_rows = [row for row in source_rows if row[6] == "Y" and row[7] == "N"]
    source_counts = {
        "lb_essential": sum(row[14] == "E" for row in protein_rows),
        "lb_nonessential": sum(row[14] == "NE" for row in protein_rows),
        "m9_essential": sum(row[19] == "E" for row in protein_rows),
        "m9_nonessential": sum(row[19] == "NE" for row in protein_rows),
    }
    observed_contract = {
        "source_rows": len(source_rows),
        "protein_coding_nonpseudo_rows": len(protein_rows),
        **source_counts,
    }
    if (
        expected_source_counts is not None
        and observed_contract != expected_source_counts
    ):
        raise DataValidationError(
            f"Choe source snapshot contract changed: {observed_contract}"
        )

    mapped_rows: dict[str, tuple[object, ...]] = {}
    unmapped: list[str] = []
    coordinate_mismatches: list[str] = []
    for row_number, row in enumerate(protein_rows, start=3):
        b_number = _text(row[5], f"row {row_number} locus tag")
        if not is_b_number(b_number):
            raise DataValidationError(
                f"row {row_number}: malformed source locus tag {b_number!r}"
            )
        _validate_source_call(row, row_number, ecipkm_index=13, call_index=14)
        _validate_source_call(row, row_number, ecipkm_index=18, call_index=19)
        if b_number not in registry.search_universe:
            unmapped.append(b_number)
            continue
        if b_number in mapped_rows:
            raise DataValidationError(f"duplicate Choe source locus tag: {b_number}")
        mapped_rows[b_number] = row
        record = registry.require(b_number)
        source_coordinates = (
            _integer(row[1], f"{b_number} start"),
            _integer(row[2], f"{b_number} end"),
        )
        if source_coordinates != (record.start, record.end):
            coordinate_mismatches.append(b_number)

    observations: list[EssentialityObservation] = []
    summaries: list[EssentialitySummary] = []
    for record in registry:
        row = mapped_rows.get(record.b_number)
        if row is None:
            summaries.append(
                EssentialitySummary(
                    b_number=record.b_number,
                    classification="unknown",
                    m9_call_raw=None,
                    lb_call_raw=None,
                    m9_ecipkm=None,
                    lb_ecipkm=None,
                    cross_condition_pattern=None,
                    evidence_conflict=False,
                    basis_observation_ids=(),
                    study_id=None,
                )
            )
            continue
        lb = _observation(record.b_number, row, "lb")
        m9 = _observation(record.b_number, row, "m9_glucose")
        observations.extend((lb, m9))
        pattern = f"LB_{lb.source_call_raw}/M9_{m9.source_call_raw}"
        if m9.source_call_raw == "E":
            classification: EssentialityClass = (
                "essential" if lb.source_call_raw == "E" else "conditionally_essential"
            )
        elif lb.source_call_raw == "E":
            classification = "ambiguous"
        else:
            classification = "nonessential"
        summaries.append(
            EssentialitySummary(
                b_number=record.b_number,
                classification=classification,
                m9_call_raw=m9.source_call_raw,
                lb_call_raw=lb.source_call_raw,
                m9_ecipkm=m9.ecipkm,
                lb_ecipkm=lb.ecipkm,
                cross_condition_pattern=pattern,
                evidence_conflict=(
                    lb.source_call_raw == "E" and m9.source_call_raw == "NE"
                ),
                basis_observation_ids=(lb.observation_id, m9.observation_id),
                study_id=_STUDY_ID,
            )
        )

    missing = tuple(sorted(registry.search_universe - set(mapped_rows)))
    summary_counts = {
        str(classification): count
        for classification, count in Counter(
            summary.classification for summary in summaries
        ).items()
    }
    return ParsedEssentiality(
        observations=tuple(observations),
        summaries=tuple(summaries),
        report=EssentialityImportReport(
            source_rows=len(source_rows),
            protein_coding_nonpseudo_rows=len(protein_rows),
            source_call_counts=source_counts,
            mapped_source_genes=len(mapped_rows),
            unmapped_source_ids=tuple(sorted(unmapped)),
            canonical_genes_without_measurement=missing,
            coordinate_mismatches=tuple(sorted(coordinate_mismatches)),
            summary_counts=summary_counts,
        ),
    )


def write_essentiality_tables(
    parsed: ParsedEssentiality,
    observation_path: Path,
    summary_path: Path,
) -> None:
    observations = pa.Table.from_pylist(
        [asdict(row) for row in parsed.observations], schema=OBSERVATION_SCHEMA
    )
    summaries = pa.Table.from_pylist(
        [row.as_arrow_row() for row in parsed.summaries], schema=SUMMARY_SCHEMA
    )
    _atomic_parquet(observations, observation_path)
    _atomic_parquet(summaries, summary_path)


def _observation(
    b_number: str, row: tuple[object, ...], assay: Literal["lb", "m9_glucose"]
) -> EssentialityObservation:
    if assay == "lb":
        offset = 10
        medium_name = "LB medium"
        medium_recipe = "Luria-Bertani rich medium"
        glucose_g_l = None
        target_relevance = "condition_control"
    else:
        offset = 15
        medium_name = "M9 minimal medium + 0.2% glucose"
        medium_recipe = "M9 minimal salts with 0.2% w/v D-glucose"
        glucose_g_l = 2.0
        target_relevance = "A"
    source_call = _text(row[offset + 4], f"{b_number} {assay} call")
    return EssentialityObservation(
        observation_id=f"{_STUDY_ID}:{b_number}:{assay}",
        b_number=b_number,
        gene_symbol_raw=_optional_text(row[0]),
        study_id=_STUDY_ID,
        assay_id=assay,
        source_call_raw=source_call,
        normalized_call="essential" if source_call == "E" else "nonessential",
        insertion_count=_integer(row[offset], f"{b_number} {assay} insertion"),
        ipkm=_number(row[offset + 1], f"{b_number} {assay} IPKM"),
        end_corrected_insertion_count=_integer(
            row[offset + 2], f"{b_number} {assay} corrected insertion"
        ),
        ecipkm=_number(row[offset + 3], f"{b_number} {assay} ecIPKM"),
        effect_metric="ecIPKM",
        threshold_rule="ecIPKM <= 2.2",
        classification_basis="author_call_validated_against_threshold",
        strain_name="Escherichia coli K-12 MG1655",
        reference_accession=REFERENCE_ACCESSION,
        medium_name=medium_name,
        medium_recipe=medium_recipe,
        glucose_g_l=glucose_g_l,
        oxygenation="aerobic",
        temperature_c=37.0,
        perturbation_type="Tn5 transposon insertion library",
        coverage_status="measured",
        target_relevance_tier=target_relevance,
        source_start=_integer(row[1], f"{b_number} start"),
        source_end=_integer(row[2], f"{b_number} end"),
        pec_call_raw=_optional_text(row[8]),
        gerdes_call_raw=_optional_text(row[9]),
        doi=_DOI,
        license_spdx="CC-BY-4.0",
    )


def _validate_source_call(
    row: tuple[object, ...], row_number: int, *, ecipkm_index: int, call_index: int
) -> None:
    ecipkm = _number(row[ecipkm_index], f"row {row_number} ecIPKM")
    call = _text(row[call_index], f"row {row_number} essentiality")
    if call not in {"E", "NE"}:
        raise DataValidationError(f"row {row_number}: unexpected call {call!r}")
    if (ecipkm <= 2.2) != (call == "E"):
        raise DataValidationError(
            f"row {row_number}: author call disagrees with ecIPKM threshold"
        )


def _text(value: object, label: str) -> str:
    if not isinstance(value, str) or not value:
        raise DataValidationError(f"{label} is not non-empty text")
    return value


def _optional_text(value: object) -> str | None:
    return value if isinstance(value, str) and value else None


def _integer(value: object, label: str) -> int:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise DataValidationError(f"{label} is not numeric")
    result = int(value)
    if result != value:
        raise DataValidationError(f"{label} is not an integer")
    return result


def _number(value: object, label: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise DataValidationError(f"{label} is not numeric")
    return float(value)


def _atomic_parquet(table: pa.Table, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent, delete=False
    ) as handle:
        temporary = Path(handle.name)
    try:
        pq.write_table(table, temporary, compression="zstd", use_dictionary=True)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)
